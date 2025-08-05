from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
from .models import *
from django.template.loader import render_to_string
from django.core.mail import EmailMultiAlternatives, send_mail
from django.contrib import messages
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.contrib.sites.shortcuts import get_current_site
from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.views.decorators.csrf import csrf_exempt
from openpyxl import load_workbook, Workbook
from io import BytesIO
from django.core.files.base import ContentFile
from django.core.mail import EmailMessage
from django.conf import settings
from django.http import JsonResponse
import stripe
import os
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch
from django.http import FileResponse
from django.conf import settings
import io
import qrcode
import json
from django.db.models import F, ExpressionWrapper, fields
from django.utils.timezone import now
from django.contrib.auth.decorators import user_passes_test, login_required


def generate_invoice(request, id):
    # Fetch Order
    order = Orders.objects.get(id=id)
    billing = BillingDetail.objects.filter(user=order.user).first()
    items = order.backlink_cart.all()

    # Buffer for PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30
    )
    elements = []
    
    # Custom Styles
    styles = getSampleStyleSheet()
    
    # Enhanced custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2E86AB'),
        spaceAfter=20,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold'
    )
    
    company_style = ParagraphStyle(
        'CompanyStyle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#333333'),
        spaceAfter=15,
        alignment=1,  # Center alignment
        leading=16
    )
    
    section_header_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2E86AB'),
        spaceBefore=15,
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#555555'),
        spaceAfter=12,
        leading=14
    )
    
    total_style = ParagraphStyle(
        'TotalStyle',
        parent=styles['Normal'],
        fontSize=16,
        textColor=colors.HexColor('#2E86AB'),
        fontName='Helvetica-Bold',
        alignment=2,  # Right alignment
        spaceBefore=15,
        spaceAfter=15
    )

    # Header with Logo and Company Info
    header_data = []
    
    # Logo (if exists)
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'digital_king.png')
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=120, height=40)
        logo.hAlign = 'LEFT'
    else:
        logo = Paragraph("<b>LOGO</b>", section_header_style)
    
    # Company info aligned right
    company_info = """
    <b>SoftSprint [ BACKLINK WORKS ]</b><br/>
    14, Kemparup Main Road<br/>
    Bengaluru 560024, KA, INDIA<br/>
    <b>Email:</b> hello@backlinkworks.com<br/>
    <b>Website:</b> backlinkworks.com
    """
    company_para = Paragraph(company_info, company_style)
    
    # Create header table
    header_table = Table([[logo, company_para]], colWidths=[200, 300])
    
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 20))


    title_table = Table([["INVOICE CUM RECEIPT"]], colWidths=[500])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#215C5C')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 18),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 25))

    # Order and Billing Info Side by Side
    order_info = f"""
    <b>Invoice Details:</b><br/>
    <b>Invoice No:</b> {order.order_id}<br/>
    <b>Invoice Date:</b> {order.created_at.strftime('%d/%m/%Y')}<br/>
    <b>Payment Status:</b> <font color="#28A745">{order.payment_status}</font><br/>
    <b>Payment Mode:</b> Online
    """
    
    billing_info = f"""
    <b>Bill To:</b><br/>
    <b>{billing.first_name} {billing.last_name}</b><br/>
    {billing.company_name}<br/>
    {billing.address}<br/>
    {billing.city}, {billing.state} - {billing.pin_code}<br/>
    <b>Email:</b> {billing.email}
    """
    
    info_table = Table([
        [Paragraph(order_info, info_style), Paragraph(billing_info, info_style)]
    ], colWidths=[250, 250])
    info_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 25))

    # Enhanced Items Table
    data = [['#', 'Service Description', 'Qty', 'Unit Price', 'Amount']]
    total_amount = 0
    
    for idx, item in enumerate(items, 1):
        amount = float(item.package_name.price)
        total_amount += amount
        data.append([
            str(idx), 
            item.package_name.title, 
            '1', 
            f"$ {amount:,.2f}",
            f"$ {amount:,.2f}"
        ])

    data.append(['', '', '', 'Subtotal:', f"$ {total_amount:,.2f}"])

    data.append(['', '', '', 'Total Amount:', f"$ {total_amount:,.2f}"])

    table = Table(data, colWidths=[30, 250, 50, 85, 85])
    table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#215C5C")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows styling
        ('FONTNAME', (0, 1), (-1, -4), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -4), 10),
        ('ALIGN', (0, 1), (0, -4), 'CENTER'),  # Serial numbers
        ('ALIGN', (2, 1), (-1, -4), 'CENTER'),  # Qty, Price, Amount
        ('ALIGN', (1, 1), (1, -4), 'LEFT'),     # Description
        
        # Subtotal and total rows
        ('FONTNAME', (0, -3), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -3), (-1, -1), 11),
        ('ALIGN', (3, -3), (-1, -1), 'RIGHT'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F8F9FA')),
        
        # Grid and borders
        ('GRID', (0, 0), (-1, -4), 1, colors.HexColor('#DEE2E6')),
        ('LINEBELOW', (0, -3), (-1, -3), 1, colors.HexColor('#DEE2E6')),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.HexColor('#2E86AB')),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        
        # Alternating row colors for better readability
        ('ROWBACKGROUNDS', (0, 1), (-1, -4), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 25))

    # Terms & Conditions with better formatting
    elements.append(Paragraph("Terms & Conditions", section_header_style))
    terms = """
    <b>1. Service Delivery:</b> Digital SEO services will be delivered as per the timeline mentioned in order confirmation.<br/><br/>
    <b>2. Service Nature:</b> All services are digital in nature and reports will be shared via email.<br/><br/>
    <b>3. Payment Terms:</b> Payment has been received online and is non-refundable after service initiation.<br/><br/>
    <b>4. Support:</b> Post-delivery support available for 30 days from completion date.
    """
    terms_para = Paragraph(terms, info_style)
    
    # Terms box with background
    terms_table = Table([[terms_para]], colWidths=[500])
    terms_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8F9FA')),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6')),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
    ]))
    elements.append(terms_table)
    elements.append(Spacer(1, 20))

    # QR Code and Footer
    footer_elements = []
    
    # Generate QR Code
    qr = qrcode.make(f"Order ID: {order.order_id}\nAmount: ₹{total_amount}")
    qr_path = os.path.join(settings.MEDIA_ROOT, f"qr_{order.order_id}.png")
    qr.save(qr_path)
    qr_image = Image(qr_path, width=80, height=80)
    
    # Footer text
    footer_text = """
    <b>Thank you for your business!</b><br/>
    For support: hello@backlinkworks.com<br/>
    Scan QR for order details
    """
    footer_para = Paragraph(footer_text, info_style)
    footer_table = Table([[qr_image, footer_para]], colWidths=[100, 400])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
    ]))
    elements.append(footer_table)

    # Add page numbers and footer
    def add_page_number(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 9)
        page_num = canvas.getPageNumber()
        text = f"Page {page_num}"
        canvas.drawRightString(570, 30, text)
        
        # Add a subtle footer line
        canvas.setStrokeColor(colors.HexColor('#DEE2E6'))
        canvas.line(30, 50, 570, 50)
        canvas.restoreState()

    # Build PDF with page template
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
    buffer.seek(0)
    
    # Clean up QR code file
    if os.path.exists(qr_path):
        os.remove(qr_path)
    return FileResponse(buffer, as_attachment=True, filename=f"invoice_{order.order_id}.pdf")
stripe.api_key = settings.STRIPE_SECRET_KEY


def index(request):
    return render(request,"index.html")


def dashboard(request):
    return render(request,"dashboard.html")


def about(request):
    return render(request,"about.html")


def contact_us(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        subject = request.POST.get('subject')
        message = request.POST.get('message')
        Contact.objects.create(
            name=name,
            email=email,
            subject=subject,
            message=message
        )

        messages.success(request, "Thank you! Your message has been sent.")
        return redirect('index')

    return render(request,"contact_us.html")


def signin(request):
    if request.method == 'POST':
        name = request.POST.get("username")
        pws = request.POST.get("password")
        user = authenticate(request, username=name, password=pws)
        if user is not None:
            login(request, user)
            if 'temp_cart_id' in request.session:
                return redirect('move_temp_to_cart')
            next_url = request.POST.get("next") or request.GET.get("next")
            if not next_url:
                next_url = "dashboard" if user.is_superuser else "my_account"

            messages.success(request, f"Welcome back to Digital King's {request.user.first_name}...!")
            return redirect(next_url)
        else:
            messages.error(request, "Invalid username or password.")
            return redirect('signin')

    return render(request, "auth/signin.html")


def signup(request):
    if request.method == 'POST':
        form = UserRegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_active = False
            user.save()
            otp_entry, _ = User_OTP.objects.get_or_create(user=user)
            otp_entry.generate_otp()

            context = {
                'otp': otp_entry.otp,
                'logo_url': 'http://127.0.0.1:8000/static/assets/img/logo.png',
            }
            html_content = render_to_string('auth/otp_mail.html', context)
            text_content = f'Your OTP is {otp_entry.otp}'
            subject = 'Please Verify Your Email Address on Digital King'
            from_email = settings.DEFAULT_FROM_EMAIL
            to_email = [user.email]
            msg = EmailMultiAlternatives(subject, text_content, from_email, to_email)
            msg.attach_alternative(html_content, "text/html")
            msg.send()
            request.session['user_id'] = user.id
            messages.info(request, 'We sent an OTP to your email.')
            return redirect('verify_otp')
        else:
            messages.error(request,form.errors)
    else:
        form = UserRegisterForm()
    return render(request,"auth/signup.html",{'form':form})


def privacy_policy(request):
    return render(request,"privacy_policy.html")


def verify_otp(request):
    user_id = request.session.get('user_id')
    if not user_id:
        return redirect('signup')
    user = User.objects.get(id=user_id)
    otp_record = User_OTP.objects.get(user=user)
    if request.method == 'POST':
        form = OTPForm(request.POST)
        if form.is_valid():
            if otp_record.is_expired():
                otp_record.delete()
                user.delete()
                messages.error(request, "OTP expired. Please request a new one.")
                return redirect('verify_otp')
            if form.cleaned_data['otp'] == otp_record.otp:
                user.is_active = True
                user.save()
                otp_record.delete()
                user.backend = 'main.authentication.EmailBackend'
                context = {
                    'name': user.first_name,
                    'email':user.email
                }
                
                html_content = render_to_string('auth/welcome_mail.html', context)
                text_content = f'Welcome to Digital King – Your Account is Ready! '
    
                subject = 'Please Verify Your Email Address on Digital King'
                from_email = settings.DEFAULT_FROM_EMAIL
                to_email = [user.email]
                
                msg = EmailMultiAlternatives(subject, text_content, from_email, to_email)
                msg.attach_alternative(html_content, "text/html")
                msg.send()
                messages.success(request, 'Registration successful and verified.')
                return redirect('signin')
            else:
                messages.error(request,'Invalid OTP')
                return redirect('verify_otp')
    else:
        form = OTPForm()
    return render(request,'auth/verify_otp.html',{'form':form})


def resend_otp(request):
    user_id = request.session.get('user_id')
    if request.user.is_authenticated:
        return redirect('home')
    if not user_id:
        messages.error(request, "Session expired. Please register again.")
        return redirect('signup')

    user = User.objects.get(id=user_id)
    otp_record, _ = User_OTP.objects.get_or_create(user=user)

    if not otp_record.is_expired() and (timezone.now() - otp_record.created).seconds < 60:
        messages.warning(request, "Please wait before requesting a new OTP.")
        return redirect('verify_otp')
    otp_record.generate_otp()
    send_mail(
        subject='Your New OTP Code',
        message=f'Your new OTP is {otp_record.otp}',
        from_email='youremail@example.com',
        recipient_list=[user.email],
    )
    messages.success(request, "A new OTP has been sent to your email.")
    return redirect('verify_otp')


def password_reset_request(request):
    if request.method == 'POST':
        form = PasswordResetRequestForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            try:
                user = User.objects.get(email=email)
                token = default_token_generator.make_token(user)
                uid = urlsafe_base64_encode(str(user.pk).encode())
                current_site = get_current_site(request)
                mail_subject = 'Password Reset Request'
                message = render_to_string('forgot/password_reset_email.html', {
                    'user': user,
                    'domain': current_site.domain,
                    'uid': uid,
                    'token': token,
                })
                send_mail(mail_subject, message, settings.DEFAULT_FROM_EMAIL, [email])
                messages.success(request, 'Password reset link has been sent to your email address.')
                return redirect('password_reset_done')
            except User.DoesNotExist:
                messages.error(request, 'Email address not found.')
                return redirect('password_reset_request')
        else:
            for field in form.errors:
                messages.error(request, form.errors[field])
    else:
        form = PasswordResetRequestForm()
    return render(request, 'forgot/password_reset_request.html', {'form': form})


def password_reset_done(request):
    return render(request, 'forgot/password_reset_done.html')



def password_reset_confirm(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = get_user_model().objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    if user and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = PasswordResetForm(user, request.POST)
            if form.is_valid():
                user.set_password(form.cleaned_data['new_password1'])
                user.save()
                messages.success(request, 'Your password has been successfully reset.')
                return redirect('password_reset_complete')
            else:
                for field in form:
                    for error in field.errors:
                        messages.error(request, error)
        else:
            form = PasswordResetForm(user)
        return render(request, 'forgot/password_reset_confirm.html', {'form': form})
    else:
        messages.error(request, 'The password reset link is invalid or has expired.')
        return redirect('password_reset_invalid')


def password_reset_complete(request):
    return render(request, 'forgot/password_reset_complete.html')


def signout(request):
    logout(request)
    return redirect('index')


# My account
def my_account(request):
    return render(request,"my_account/my_account.html")



def package(request, package_name):
    package = Package.objects.get(title=package_name)

    if not request.session.session_key:
        request.session.create()

    if request.method == 'POST':
        website_url = request.POST.get('website_url')
        keywords = [request.POST.get(f'keyword_{i}') for i in range(1, 11)]
        image_url = request.POST.get('image_url')
        youtube_url = request.POST.get('youtube_url')
        article_document = request.FILES.get('article_document')

        if request.user.is_authenticated:
            Backlink_cart.objects.create(
                user=request.user,
                package_name=package,
                website_url=website_url,
                keyword_1=keywords[0],
                keyword_2=keywords[1],
                keyword_3=keywords[2],
                keyword_4=keywords[3],
                keyword_5=keywords[4],
                keyword_6=keywords[5],
                keyword_7=keywords[6],
                keyword_8=keywords[7],
                keyword_9=keywords[8],
                keyword_10=keywords[9],
                image_url=image_url,
                youtube_url=youtube_url,
                article_document=article_document
            )
            messages.success(request, "Backlink added to your cart successfully!")
            return redirect('my_cart')
        else:
            temp_cart = TempBacklinkCart.objects.create(
                session_key=request.session.session_key,
                package_name=package,
                website_url=website_url,
                keyword_1=keywords[0],
                keyword_2=keywords[1],
                keyword_3=keywords[2],
                keyword_4=keywords[3],
                keyword_5=keywords[4],
                keyword_6=keywords[5],
                keyword_7=keywords[6],
                keyword_8=keywords[7],
                keyword_9=keywords[8],
                keyword_10=keywords[9], 
                image_url=image_url,
                youtube_url=youtube_url,
                article_document=article_document
            )
            request.session['temp_cart_id'] = temp_cart.id
            messages.info(request, "Please login or register to continue.")
            return redirect('signin')

    return render(request, "package.html", {
        'package_name': package_name,
        'range': range(1, 11),
        'package_details': package
    })


@login_required
def move_temp_to_cart(request):
    temp_cart_id = request.session.get('temp_cart_id')
    if temp_cart_id:
        temp_item = TempBacklinkCart.objects.filter(id=temp_cart_id).first()
        if temp_item:
            Backlink_cart.objects.create(
                user=request.user,
                package_name=temp_item.package_name,
                website_url=temp_item.website_url,
                keyword_1=temp_item.keyword_1,
                keyword_2=temp_item.keyword_2,
                keyword_3=temp_item.keyword_3,
                keyword_4=temp_item.keyword_4,
                keyword_5=temp_item.keyword_5,
                keyword_6=temp_item.keyword_6,
                keyword_7=temp_item.keyword_7,
                keyword_8=temp_item.keyword_8,
                keyword_9=temp_item.keyword_9,
                keyword_10=temp_item.keyword_10,
                image_url=temp_item.image_url,
                youtube_url=temp_item.youtube_url,
                article_document=temp_item.article_document
            )
            temp_item.delete()
            del request.session['temp_cart_id']
    return redirect('my_cart')
    
def my_cart(request):
    backlink_items = Backlink_cart.objects.filter(user=request.user, is_paid=False)
    total_price = backlink_items.aggregate(
        total=Sum('package_name__price')
    )['total'] or 0
    return render(request,"my_account/my_cart.html",{'backlink_items':backlink_items,'total_price':total_price})



def delete_cart(request,item_id):
    item = Backlink_cart.objects.get(id=item_id)
    item.delete()
    return redirect('my_cart')


@csrf_exempt
def stripe_webhook(request):
    payload = request.body
    sig_header = request.META['HTTP_STRIPE_SIGNATURE']
    endpoint_secret = "your_webhook_secret"
    try:
        event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
    except ValueError:
        return JsonResponse({'error': 'Invalid payload'}, status=400)
    except stripe.error.SignatureVerificationError:
        return JsonResponse({'error': 'Invalid signature'}, status=400)
    if event['type'] == 'payment_intent.succeeded':
        payment_intent = event['data']['object']
        order_id = payment_intent['metadata']['order_id']
        Orders.objects.filter(id=order_id).update(payment_status="Completed")
    return JsonResponse({'status': 'success'})



def checkout(request):
    backlink_items = Backlink_cart.objects.filter(user=request.user, is_paid=False)
    total_price = backlink_items.aggregate(
        total=Sum('package_name__price')
    )['total'] or 0

    try:
        billing_instance = BillingDetail.objects.get(user=request.user)
    except BillingDetail.DoesNotExist:
        billing_instance = None

    if request.method == 'POST':
        if billing_instance:
            form = BillingDetailForm(request.POST, instance=billing_instance)
        else:
            form = BillingDetailForm(request.POST)

        if form.is_valid():
            billing_detail = form.save(commit=False)
            billing_detail.user = request.user
            billing_detail.save()

            # Create Order
            order = Orders.objects.create(
                user=request.user,
                payment_status="Pending",
                total_amount=total_price
            )
            order.backlink_cart.set(backlink_items)

            # Redirect to payment page
            return redirect('payment_page', order_id=order.id)

    else:
        if billing_instance:
            form = BillingDetailForm(instance=billing_instance)
        else:
            initial_data = {
                'email': request.user.email,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
            }
            form = BillingDetailForm(initial=initial_data)

    return render(request, "my_account/checkout.html", {
        'total_price': total_price,
        'form': form
    })

def payment_page(request, order_id):
    order = Orders.objects.get(id=order_id)
    total_price_cents = int(float(order.total_amount) * 100)

    if request.method == 'POST':
        # Create Payment Intent
        intent = stripe.PaymentIntent.create(
            amount=total_price_cents,
            currency='usd',
            metadata={'order_id': order.id},
        )
        return JsonResponse({'client_secret': intent.client_secret})

    return render(request, "my_account/payment.html", {
        'order': order,
        'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY
    })
    
    
@csrf_exempt
def update_order_status(request):
    data = json.loads(request.body)
    order_id = data.get('order_id')
    order = Orders.objects.get(id=order_id)
    order.payment_status = 'Completed'
    order.work_status = 'onprogress'
    order.save()
    order.backlink_cart.update(is_paid=True)
    return JsonResponse({'status': 'success'})

def payment_success(request):
    return render(request, "my_account/payment_success.html")

def payment_cancel(request):
    return render(request, "my_account/payment_cancel.html")


def edit_billing_address(request):
    billing_instance = BillingDetail.objects.get(user=request.user)
    if request.method == 'POST':
        form = BillingDetailForm(request.POST, instance=billing_instance)
        if form.is_valid():
            billing_detail = form.save(commit=False)
            billing_detail.user = request.user
            billing_detail.save()
            return redirect('user_ billing_address')
    else:
        form = BillingDetailForm(instance=billing_instance)
    return render(request,"my_account/edit_billing_address.html",{'form':form})


def orders(request):
    orders = Orders.objects.filter(user=request.user)
    return render(request,"my_account/orders.html",{'orders':orders})


def view_order(request,ord_id):
    order = Orders.objects.get(order_id=ord_id)
    return render(request,"my_account/view_order.html",{'order':order})


def downloads(request):
    downloads = Orders.objects.filter(payment_status="Completed")
    return render(request,"my_account/downloads.html",{'downloads':downloads})


def user_billing_address(request):
    try:
        billing_detail = BillingDetail.objects.get(user=request.user)
    except:
        billing_detail = None
    return render(request,"my_account/billing_address.html",{'billing_detail':billing_detail})

def user_customer_query(request):
    if request.method == 'POST':
        form = CustomerQueryForm(request.POST, request.FILES)
        if form.is_valid():
            query = form.save(commit=False)
            query.user = request.user
            query.save()
            return redirect('customer_query')
    else:
        form = CustomerQueryForm()
    return render(request,"my_account/customer_query.html",{'form':form})

def account_details(request):
    return render(request,"my_account/account_details.html")


# Admin Dashboard

@user_passes_test(lambda u:u.is_superuser)
def dashboard(request):
    if User.objects.exists():
        last_user = User.objects.latest('date_joined').date_joined.strftime('%b %d, %Y, %I:%M %p')
    else:
        last_user = "No users"
    user_count = User.objects.count()
    if Orders.objects.exists():
        last_order = Orders.objects.latest('created_at').created_at.strftime('%b %d, %Y')
    else:
        "No orders"
    order_count = Orders.objects.filter(payment_status="Completed").count()
    pending_order_count = Orders.objects.filter(payment_status="Completed").filter(work_status="onprogress").count()
    upcoming_orders = Backlink_cart.objects.annotate(
    upcoming_delivery_date=ExpressionWrapper(
        F('created_at') + F('package_name__days_to_complete') * 86400,
        output_field=fields.DateTimeField()
    )
    ).filter(upcoming_delivery_date__gt=now()).order_by('upcoming_delivery_date')[0]
    return render(request,"admin_dashboard/dashboard.html",{'last_user':last_user,'user_count':user_count,'last_order':last_order,'order_count':order_count,'pending_order_count':pending_order_count,'upcoming_orders':upcoming_orders})



def order_dashboard(request):
    total_orders = Orders.objects.filter(payment_status="Completed").count()
    completed_orders = Orders.objects.filter(payment_status="Completed").filter(work_status="Completed").count()
    print("========",completed_orders)
    pending_orders = Orders.objects.filter(payment_status="Completed").filter(work_status="onprogress").count()
    scheduled_orders = Orders.objects.filter(payment_status='Completed').filter(work_status="onprogress").order_by('order_date')
    return render(request,"admin_dashboard/order_dashboard.html",{'total_orders':total_orders,'completed_orders':completed_orders,'pending_orders':pending_orders,'scheduled_orders':scheduled_orders})


def orders_list(request):
    orders = Orders.objects.filter(payment_status="Completed")
    return render(request,"admin_dashboard/orders_list.html",{'orders':orders})


def admin_view_order(request,id):
    order = Orders.objects.get(id=id)
    return render(request,"admin_dashboard/admin_view_order.html",{'order':order})


def customer_query(request):
    querys = CustomerQuery.objects.all()
    return render(request,"admin_dashboard/customer_query.html",{'querys':querys})


def upload_report(request, id):
    order = get_object_or_404(Orders, id=id)

    if request.method == 'POST':
        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)

        for item in order.backlink_cart.all():
            file = request.FILES.get(f'report_{item.id}')
            if file:
                uploaded_wb = load_workbook(file)
                source_sheet = uploaded_wb.active
                new_sheet = merged_wb.create_sheet(title=str(item.package_name)[:31])
                for row in source_sheet.iter_rows(values_only=True):
                    new_sheet.append(row)

        # Save the workbook to memory
        excel_stream = BytesIO()
        merged_wb.save(excel_stream)
        excel_stream.seek(0)

        # Save to model
        filename = f"order_{order.order_id}_report.xlsx"
        order.report_file.save(filename, ContentFile(excel_stream.getvalue()))
        order.work_status = "Completed"
        order.save()

        # Prepare and send email with attachment
        subject = f"Report for Your Order #{order.order_id} is Ready"
        message = render_to_string('email/report_ready.html', {
            'user': order.user,
            'order': order,
        })

        email = EmailMessage(
            subject=subject,
            body=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[order.user.email]
        )
        email.content_subtype = "html"
        
        # Attach the Excel file
        email.attach(filename, excel_stream.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        email.send()

        messages.success(request, "Report updated and sent to user via email.")
        return redirect('orders_list')

    return render(request, "admin_dashboard/upload_report.html", {'order': order})


def users_list(request):
    user_data = User.objects.all()
    return render(request,"admin_dashboard/users_list.html",{'user_data':user_data})

def billing_address(request):
    addresses = BillingDetail.objects.all()
    return render(request,"admin_dashboard/billing_address.html",{'addresses':addresses})


def contact_list(request):
    contact = Contact.objects.all()
    return render(request,"admin_dashboard/contact_list.html",{'contact':contact})

